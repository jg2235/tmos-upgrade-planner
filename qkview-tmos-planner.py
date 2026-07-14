#!/usr/bin/env python3
"""
tmos_rn_advisor.py — BIG-IP QKView config vs TMOS Release Notes association engine.

Reads a BIG-IP configuration (live from iHealth QKView via Files API, or offline
from local bigip.conf / gtm.conf), parses one or more TMOS Release Notes HTML
files, and produces a consolidated HTML + Excel report associating every
configuration element (feature-level and named-instance-level) with the
bugfixes, vulnerabilities, behavior changes, and known issues in the target
releases. Each association carries a deterministic 0-100 accuracy score.

Scoring rubric (deterministic, auditable):
  85-100  strong feature term matched in the item's Conditions text
  60-84   strong feature term matched in Symptoms / Title only
  30-59   component-level match only (config uses the item's component)
  1-29    weak lexical hits only
  0       no association (listed in Unmatched section)

Security posture:
  - Credentials only via env vars (IHEALTH_CLIENT_ID / IHEALTH_CLIENT_SECRET).
  - TLS verification always on; no subprocess, no shell, no eval.
  - Fully offline matching — config data never leaves the host.
"""

from __future__ import annotations

import argparse
import base64
import html as html_mod
import json
import os
import re
import sys
import time
from collections import OrderedDict, defaultdict
from dataclasses import dataclass, field
from datetime import datetime, timezone
from typing import Any, Dict, List, Optional, Set, Tuple

try:
    import requests
except ImportError:          # offline mode works without requests
    requests = None          # type: ignore

__version__ = "1.0.0"

# =============================================================================
# iHealth API client  (patterns reused from jg2235/BIG-IP_iHealth_QKview_Parser)
# =============================================================================

API_BASE = os.environ.get("IHEALTH_API_BASE",
                          "https://ihealth-api.f5.com/qkview-analyzer/api")
TOKEN_URL = os.environ.get(
    "IHEALTH_TOKEN_URL",
    "https://identity.account.f5.com/oauth2/ausp95ykc80HOU7SQ357/v1/token",
)
ACCEPT_JSON = "application/vnd.f5.ihealth.api+json"
ACCEPT_OCTET = "application/octet-stream"
USER_AGENT = os.environ.get("IHEALTH_UA", f"tmos-rn-advisor/{__version__}")

CONFIG_FILES = OrderedDict([        # QKView paths pulled via Files API
    ("bigip.conf",      "/config/bigip.conf"),
    ("gtm.conf",        "/config/bigip_gtm.conf"),
    ("bigip_base.conf", "/config/bigip_base.conf"),
])


class IHealthError(RuntimeError):
    pass


class IHealthClient:
    """Minimal, hardened iHealth REST client (auth + files only)."""

    def __init__(self) -> None:
        if requests is None:
            raise IHealthError("python 'requests' is required for live mode")
        self.s = requests.Session()
        self.s.headers["User-Agent"] = USER_AGENT
        self._token: Optional[str] = os.environ.get("IHEALTH_TOKEN")
        self._acquired = time.time() if self._token else 0.0

    def authenticate(self) -> None:
        cid = os.environ.get("IHEALTH_CLIENT_ID")
        sec = os.environ.get("IHEALTH_CLIENT_SECRET")
        if not (cid and sec):
            raise IHealthError("Set IHEALTH_CLIENT_ID and IHEALTH_CLIENT_SECRET")
        basic = base64.b64encode(f"{cid}:{sec}".encode()).decode()
        r = requests.post(TOKEN_URL, headers={
            "accept": "application/json",
            "authorization": f"Basic {basic}",
            "cache-control": "no-cache",
            "content-type": "application/x-www-form-urlencoded",
        }, data="grant_type=client_credentials&scope=ihealth", timeout=30)
        r.raise_for_status()
        self._token = r.json().get("access_token")
        if not self._token:
            raise IHealthError(f"No access_token in token response")
        self._acquired = time.time()

    def _auth(self) -> Dict[str, str]:
        if not self._token or (time.time() - self._acquired) > 1500:
            self.authenticate()
        return {"Authorization": f"Bearer {self._token}"}

    def _get(self, path: str, accept: str = ACCEPT_JSON,
             max_tries: int = 12, sleep_s: float = 10.0):
        url = f"{API_BASE}{path}"
        backoff = 3.0
        r = None
        for attempt in range(1, max_tries + 1):
            try:
                r = self.s.get(url, headers={"Accept": accept, **self._auth()},
                               allow_redirects=False, timeout=180)
            except (requests.exceptions.ConnectionError,
                    requests.exceptions.Timeout) as e:
                if attempt < max_tries:
                    print(f"  net err ({attempt}/{max_tries}): {e}; "
                          f"retry in {backoff:.0f}s", file=sys.stderr)
                    time.sleep(backoff); backoff = min(backoff * 2, 60); continue
                raise
            if r.status_code == 202 and attempt < max_tries:
                print(f"  202 processing ({attempt}/{max_tries}); "
                      f"waiting {sleep_s}s", file=sys.stderr)
                time.sleep(sleep_s); continue
            if r.status_code in (500, 502, 503, 504) and attempt < max_tries:
                print(f"  HTTP {r.status_code} transient ({attempt}/{max_tries}); "
                      f"retry in {backoff:.0f}s", file=sys.stderr)
                time.sleep(backoff); backoff = min(backoff * 2, 60); continue
            return r
        return r

    def list_files(self, qid: str) -> List[Tuple[str, str]]:
        r = self._get(f"/qkviews/{qid}/files")
        if r.status_code != 200:
            raise IHealthError(f"files {qid}: HTTP {r.status_code}")
        body = r.json()
        out: List[Tuple[str, str]] = []
        if isinstance(body, list):
            items = body
        else:
            items = []
            for v in body.values():
                if isinstance(v, list):
                    items.extend(v)
        for it in items:
            if isinstance(it, dict):
                fid = it.get("id") or it.get("@id")
                path = (it.get("value") or it.get("path") or it.get("name")
                        or it.get("$") or it.get("#text"))
                if fid and path:
                    out.append((str(fid), str(path)))
        return out

    def fetch_text(self, qid: str, file_hash: str) -> Optional[str]:
        r = self._get(f"/qkviews/{qid}/files/{file_hash}", accept=ACCEPT_OCTET)
        if r.status_code != 200:
            return None
        return r.content.decode("utf-8", errors="replace")

    def identity(self, qid: str) -> Dict[str, str]:
        """Pull hostname/version from diagnostics system_information."""
        r = self._get(f"/qkviews/{qid}/diagnostics.json?set=hit")
        info: Dict[str, str] = {}
        if r.status_code == 200:
            try:
                body = r.json()
                def walk(o):
                    if isinstance(o, dict):
                        for k, v in o.items():
                            yield k, v
                            yield from walk(v)
                    elif isinstance(o, list):
                        for x in o:
                            yield from walk(x)
                for k, v in walk(body):
                    if k in ("hostname", "version", "platform", "product") \
                            and isinstance(v, (str, int)) and k not in info:
                        info[k] = str(v)
            except (ValueError, KeyError):
                pass
        return info


def fetch_config_from_ihealth(qid: str) -> Tuple[Dict[str, str], Dict[str, str]]:
    """Returns ({name: text}, identity) for the QKView's config files."""
    client = IHealthClient()
    files = client.list_files(qid)
    by_path = {p: h for h, p in files}
    out: Dict[str, str] = {}
    for name, path in CONFIG_FILES.items():
        h = by_path.get(path)
        if not h:                              # tolerate path variants
            cand = [p for p in by_path if p.endswith("/" + os.path.basename(path))]
            h = by_path.get(cand[0]) if cand else None
        if h:
            print(f"  fetching {path} ...", file=sys.stderr)
            txt = client.fetch_text(qid, h)
            if txt:
                out[name] = txt
        else:
            print(f"  {path} not present in QKView (skipped)", file=sys.stderr)
    ident = client.identity(qid)
    return out, ident


# =============================================================================
# tmsh configuration parser (brace-aware)
# =============================================================================

@dataclass
class ConfigObject:
    module: str          # ltm / gtm / apm / asm / security / sys / net ...
    otype: str           # "virtual", "pool", "profile client-ssl", "wideip a" ...
    name: str            # /Common/vs_app1_https
    body: str            # raw stanza body

    @property
    def kind(self) -> str:
        return f"{self.module} {self.otype}"


_STANZA_HEAD = re.compile(
    r"^(ltm|gtm|apm|asm|security|sys|net|auth|cm|wom|pem|ilx|vcmp)\s+(.*?)\s*\{\s*$")


def parse_tmsh(text: str) -> List[ConfigObject]:
    """Parse top-level tmsh stanzas into ConfigObjects (brace-depth tracked)."""
    objs: List[ConfigObject] = []
    lines = text.splitlines()
    i, n = 0, len(lines)
    while i < n:
        line = lines[i]
        m = _STANZA_HEAD.match(line.strip()) if not line.startswith((" ", "\t")) else None
        if m:
            module = m.group(1)
            rest = m.group(2)
            depth = line.count("{") - line.count("}")
            body_lines: List[str] = []
            i += 1
            while i < n and depth > 0:
                body_lines.append(lines[i])
                depth += lines[i].count("{") - lines[i].count("}")
                i += 1
            # split "rest" into otype + name: name = last token starting with /
            # or the last token; otype = everything before it
            tokens = rest.split()
            name_idx = None
            for idx in range(len(tokens) - 1, -1, -1):
                if tokens[idx].startswith("/"):
                    name_idx = idx
                    break
            if name_idx is None:
                name_idx = len(tokens) - 1
            otype = " ".join(tokens[:name_idx]) or tokens[name_idx]
            name = tokens[name_idx]
            objs.append(ConfigObject(module, otype, name, "\n".join(body_lines)))
        else:
            i += 1
    return objs


# =============================================================================
# Feature taxonomy — config features → RN search terms
# =============================================================================
# strong terms: matched (case-insensitive regex) against Conditions =>
#               85-100 band; against Symptoms/Title => 60-84 band.
# weak terms:   generic lexical support; alone => 1-29 band.

@dataclass
class Feature:
    fid: str
    label: str
    components: Set[str]                 # canonical RN Component names
    strong: List[str]                    # regex fragments
    weak: List[str] = field(default_factory=list)
    instances: List[str] = field(default_factory=list)   # named config objects
    detail: str = ""                     # human note on why detected


COMP_LTM = "Local Traffic Manager"
COMP_GTM = "Global Traffic Manager (DNS)"
COMP_TMOS = "TMOS"
COMP_ASM = "Application Security Manager"
COMP_APM = "Access Policy Manager"
COMP_AFM = "Advanced Firewall Manager"
COMP_BOT = "Bot Defense"
COMP_ITMM = "In-tmm monitors"
COMP_SSLO = "SSL Orchestrator"
COMP_PI = "Protocol Inspection"
COMP_DM = "Device Management"

# words too generic to ever count as strong
_GENERIC_WEAK = ["tmm", "config", "traffic"]


def _feature_defs() -> List[Feature]:
    """Static taxonomy. Detection populates .instances from parsed config."""
    F = Feature
    return [
        F("virtual-server", "LTM Virtual Servers", {COMP_LTM},
          strong=[r"virtual server", r"\bvirtual-server\b", r"\bvip\b"],
          weak=[r"listener", r"destination address"]),
        F("pool-lb", "LTM Pools / load balancing", {COMP_LTM},
          strong=[r"\bpool member", r"\bpools?\b.*load.balanc", r"load.balancing",
                  r"least.connections", r"round.robin", r"slow.ramp", r"priority.group"],
          weak=[r"\bpool\b", r"\bmember\b", r"\bnode\b"]),
        F("health-monitor", "Health monitors (HTTP/HTTPS)", {COMP_LTM, COMP_ITMM},
          strong=[r"health monitor", r"\bmonitor(s|ing)?\b.*(http|https|probe)",
                  r"in.tmm monitor", r"\bbigd\b", r"monitor probe"],
          weak=[r"\bmonitor\b", r"marked down", r"\bprobe\b"]),
        F("client-ssl", "Client SSL / TLS termination", {COMP_LTM, COMP_TMOS},
          strong=[r"client.?ssl", r"ssl profile", r"tls 1\.[23]", r"tlsv1\.[23]",
                  r"ssl handshake", r"cipher.group", r"\bciphers?\b",
                  r"cert.key.chain", r"client hello", r"\bsni\b", r"\balpn\b"],
          weak=[r"\bssl\b", r"\btls\b", r"handshake", r"certificate", r"\bkey\b"]),
        F("server-ssl", "Server SSL / re-encryption (mTLS to origin)", {COMP_LTM, COMP_TMOS},
          strong=[r"server.?ssl", r"serverside ssl", r"re-?encrypt",
                  r"peer.cert", r"mutual tls", r"\bmtls\b", r"trusted.ca"],
          weak=[r"\bssl\b", r"\btls\b", r"certificate"]),
        F("ocsp-stapling", "OCSP stapling", {COMP_LTM, COMP_TMOS},
          strong=[r"ocsp.stapl", r"\bocsp\b"],
          weak=[r"revocation", r"certificate status"]),
        F("cert-mgmt", "Certificates / keys on box", {COMP_TMOS, COMP_LTM},
          strong=[r"certificate (expir|renew|install|bundle|chain)",
                  r"\bcert.manager\b", r"\bcrl\b", r"ca.bundle"],
          weak=[r"certificate", r"x509"]),
        F("http-profile", "HTTP profiles (L7 proxy)", {COMP_LTM},
          strong=[r"http profile", r"\bhttp request(s)?\b", r"http header",
                  r"\bhsts\b", r"x-?forwarded-?for", r"\bxff\b", r"chunk(ed|ing)",
                  r"http response", r"\bhttp/1\.[01]\b", r"pipelin"],
          weak=[r"\bhttp\b", r"header", r"\buri\b", r"\bhost\b"]),
        F("http2", "HTTP/2 profiles", {COMP_LTM},
          strong=[r"http/?2", r"\bh2\b", r"concurrent.streams", r"\bhpack\b",
                  r"stream (reset|priorit)", r"goaway"],
          weak=[r"\bstream\b", r"multiplex"]),
        F("websocket", "WebSocket profiles", {COMP_LTM},
          strong=[r"websocket", r"\bws://", r"upgrade header"],
          weak=[r"masking"]),
        F("oneconnect", "OneConnect (connection reuse)", {COMP_LTM},
          strong=[r"one.?connect", r"connection reuse", r"\breuse\b.*connection"],
          weak=[r"keep-?alive", r"idle timeout"]),
        F("tcp-profile", "TCP profiles / congestion control", {COMP_LTM, COMP_TMOS},
          strong=[r"tcp profile", r"congestion control", r"\bbbr\b", r"\bnagle\b",
                  r"tcp (retransmit|window|timestamp|syn)", r"\bmss\b"],
          weak=[r"\btcp\b", r"\bsyn\b", r"\brst\b", r"connection"]),
        F("udp-profile", "UDP virtual servers / profiles", {COMP_LTM},
          strong=[r"udp profile", r"\budp\b.*(virtual|datagram|flow)"],
          weak=[r"\budp\b", r"datagram"]),
        F("ltm-policy", "LTM traffic policies (L7 policy)", {COMP_LTM},
          strong=[r"(ltm|traffic|local traffic) polic(y|ies)", r"policy rule",
                  r"first-match", r"\bcentralized polic"],
          weak=[r"\bpolicy\b", r"\brule\b"]),
        F("irule", "iRules (TCL)", {COMP_LTM},
          strong=[r"irules?\b", r"\btcl\b", r"HTTP::", r"DNS::", r"SSL::",
                  r"when (HTTP|CLIENT|SERVER|DNS|LB)_"],
          weak=[r"\bscript\b", r"\bevent\b"]),
        F("persist-cookie", "Cookie persistence (incl. encryption)", {COMP_LTM},
          strong=[r"cookie persist", r"persistence cookie", r"cookie.encrypt",
                  r"bigip cookie"],
          weak=[r"\bcookie\b", r"persist"]),
        F("persist-srcaddr", "Source-address persistence", {COMP_LTM},
          strong=[r"source.?addr(ess)? persist", r"persistence record"],
          weak=[r"persist"]),
        F("snat", "SNAT / SNAT pools / automap", {COMP_LTM},
          strong=[r"\bsnat\b", r"snat.?pool", r"automap", r"source.address.translation",
                  r"port exhaustion"],
          weak=[r"\bnat\b", r"translation"]),
        F("dns-profile", "DNS profiles / DNS Express / cache / rapid-response",
          {COMP_GTM, COMP_LTM},
          strong=[r"dns (profile|cache|express|query|response|resolution)",
                  r"dns.?express", r"rapid.response", r"\bdns64\b", r"unhandled.query"],
          weak=[r"\bdns\b", r"\bquery\b", r"resolver"]),
        F("dnssec", "DNSSEC signing (keys/zones)", {COMP_GTM},
          strong=[r"dnssec", r"\bzsk\b", r"\bksk\b", r"key rollover", r"\bdnskey\b",
                  r"\brrsig\b", r"\bnsec3?\b", r"zone sign"],
          weak=[r"\bkey\b", r"\bzone\b", r"signature"]),
        F("local-bind", "Local BIND / named / ZoneRunner", {COMP_GTM, COMP_TMOS},
          strong=[r"\bbind\b", r"\bnamed\b", r"zonerunner", r"zone transfer",
                  r"\baxfr\b", r"\bixfr\b"],
          weak=[r"\bzone\b"]),
        F("gtm-wideip", "GTM/DNS Wide IPs (GSLB)", {COMP_GTM},
          strong=[r"wide.?ip", r"\bgslb\b", r"\bwip\b", r"global.availab",
                  r"topology (record|load.balanc|region)", r"\bldns\b",
                  r"persist.cidr"],
          weak=[r"\bgtm\b", r"big-?ip dns"]),
        F("gtm-pool-server", "GTM pools / servers / datacenters", {COMP_GTM},
          strong=[r"gtm (pool|server|monitor)", r"\bdatacenter\b", r"\bbig3d\b",
                  r"\bgtmd\b", r"iquery", r"virtual.server discovery",
                  r"gtm sync", r"sync group"],
          weak=[r"\bpool\b", r"\bserver\b", r"\bprober\b"]),
        F("gtm-listener", "GTM/DNS listeners", {COMP_GTM},
          strong=[r"\blistener(s)?\b"],
          weak=[r"port 53"]),
        F("asm", "ASM / WAF security policies", {COMP_ASM},
          strong=[r"\basm\b", r"\bwaf\b", r"security policy", r"attack signature",
                  r"policy builder", r"violation", r"\bbd\b daemon", r"\bbd_agent\b",
                  r"enforcement mode", r"blocking (mode|page)"],
          weak=[r"\bsecurity\b", r"signature"]),
        F("apm", "APM access profiles / policies", {COMP_APM},
          strong=[r"\bapm\b", r"access (profile|policy|session)", r"\bsso\b",
                  r"per-request policy", r"\bsaml\b", r"\boauth\b", r"\bportal access\b",
                  r"\bvpn\b", r"network access", r"session variable", r"\bapmd\b",
                  r"\bwebtop\b", r"\bldap\b.*auth", r"\bkerberos\b"],
          weak=[r"\bsession\b", r"authentication", r"\blogon\b"]),
        F("afm", "AFM firewall policies / DoS", {COMP_AFM},
          strong=[r"\bafm\b", r"firewall (rule|policy)", r"\bdos\b (profile|protect|vector)",
                  r"ddos", r"packet filter", r"ip intelligence", r"port misuse",
                  r"flow.?spec"],
          weak=[r"firewall", r"\bacl\b"]),
        F("bot-defense", "Bot Defense profiles", {COMP_BOT, COMP_ASM},
          strong=[r"bot.?defen[cs]e", r"\bbot\b (signature|profile|detect)",
                  r"browser challenge"],
          weak=[r"\bbot\b", r"challenge"]),
        F("security-logging", "Security logging profiles", {COMP_ASM, COMP_AFM, COMP_TMOS},
          strong=[r"(security-)?log(ging)? profile", r"remote logging", r"\bhsl\b",
                  r"high.speed logging", r"\bsyslog\b"],
          weak=[r"\blog(s|ging)?\b"]),
        F("mgmt-plane", "Management plane (GUI/TMUI, REST, tmsh)", {COMP_TMOS, COMP_DM},
          strong=[r"\btmui\b", r"configuration utility", r"\birest\b", r"icontrol rest",
                  r"\brestjavad\b", r"\brestnoded\b", r"\btmsh\b", r"\bmcpd\b",
                  r"config(uration)? (load|save|sync)", r"\bhttpd\b", r"management interface"],
          weak=[r"\bgui\b", r"\brest\b", r"\badmin\b", r"webui"]),
        F("ha-sync", "HA / device groups / config sync / failover", {COMP_TMOS},
          strong=[r"config.?sync", r"device group", r"failover", r"\bha\b (pair|group|state)",
                  r"traffic.group", r"mirror(ing|ed)?\b.*(connection|persist)",
                  r"\bsod\b", r"active.standby"],
          weak=[r"\bsync\b", r"standby", r"\bpeer\b"]),
        F("upgrade-install", "Software install / upgrade path", {COMP_TMOS},
          strong=[r"\bliveinstall\b", r"boot location", r"software installation",
                  r"rolling forward", r"\bucs\b (restore|load)",
                  r"fail(s|ure)? to (load|boot) after upgrad"],
          weak=[r"upgrade", r"install", r"reboot", r"migration"]),
    ]


# canonical component set always present on a running BIG-IP
_BASELINE_COMPONENTS = {COMP_TMOS, COMP_DM}

# sys provision module -> canonical RN components enabled by that module
_PROVISION_COMPONENT_MAP: Dict[str, Set[str]] = {
    "ltm": {COMP_LTM}, "gtm": {COMP_GTM}, "apm": {COMP_APM},
    "asm": {COMP_ASM, COMP_BOT}, "afm": {COMP_AFM},
    "dos": {COMP_AFM}, "sslo": {COMP_SSLO},
}
# config module -> provisioning module(s) that must be active for its
# objects to count as "in use"
_MODULE_PROVISION_GATE: Dict[str, Set[str]] = {
    "apm": {"apm"}, "asm": {"asm"},
}
_OTYPE_PROVISION_GATE: Dict[str, Set[str]] = {
    "firewall": {"afm", "dos"}, "bot-defense": {"asm"}, "dos": {"afm", "dos"},
}

# default/system objects shipped in bigip.conf on every box — never evidence
# that a module is actually in use
_DEFAULT_OBJECT_NAMES = {
    "/Common/access", "/Common/dos", "/Common/bot-defense",
    "/Common/bot-defense-device-id-generate-before-access",
    "/Common/dos-device-config", "/Common/global-network",
    "/Common/servicediscovery", "/Common/f5-default",
}
_SYSTEM_SINGLETON_OTYPES = {
    "firewall config-change-log", "firewall management-ip-rules",
    "firewall current-state", "shared-objects", "device-id",
    "log profile", "resource sandbox", "epsec",
}


def parse_provisioning(objs: List[ConfigObject]) -> Optional[Dict[str, str]]:
    """Return {module: level} from 'sys provision' stanzas, or None if the
    config set contains no provisioning info (e.g. bigip_base.conf absent)."""
    prov: Dict[str, str] = {}
    for o in objs:
        if o.module == "sys" and o.otype.startswith("provision"):
            # stanza splitter puts the module token in .name for
            # "sys provision <mod>" (no /partition prefix)
            mod = (o.name if not o.name.startswith("/")
                   else o.otype.split()[-1]).lower()
            m = re.search(r"level\s+(\S+)", o.body)
            prov[mod] = m.group(1) if m else "none"
    return prov or None


def _is_default_object(o: ConfigObject) -> bool:
    if o.name in _DEFAULT_OBJECT_NAMES:
        return True
    ot = o.otype
    for sing in _SYSTEM_SINGLETON_OTYPES:
        if ot.endswith(sing) or ot == sing:
            return True
    return False


def detect_features(objs: List[ConfigObject]
                    ) -> Tuple[List[Feature], Set[str], Dict[str, Any]]:
    """Populate feature instances from parsed config.
    Returns (features, components, provisioning_info).

    Module presence is gated on 'sys provision' when available (authoritative);
    default/system objects shipped in every bigip.conf (e.g. /Common/access,
    /Common/bot-defense, /Common/dos, firewall singletons) are never counted
    as evidence a module is in use."""
    feats = {f.fid: f for f in _feature_defs()}
    comps: Set[str] = set(_BASELINE_COMPONENTS)

    prov = parse_provisioning(objs)
    provisioned: Optional[Set[str]] = None
    if prov is not None:
        provisioned = {m for m, lvl in prov.items() if lvl.lower() != "none"}

    def module_active(gates: Set[str]) -> bool:
        if provisioned is None:
            return True          # no provisioning info -> denylist-only mode
        return bool(gates & provisioned)

    skipped: Dict[str, int] = defaultdict(int)

    def gated_out(o: ConfigObject) -> bool:
        """True if this object must NOT count (default object or module
        unprovisioned)."""
        if _is_default_object(o):
            skipped[f"default:{o.module} {o.otype}"] += 1
            return True
        gates = _MODULE_PROVISION_GATE.get(o.module)
        if gates is None and o.module == "security":
            for key, g in _OTYPE_PROVISION_GATE.items():
                if o.otype.startswith(key):
                    gates = g
                    break
        if gates and not module_active(gates):
            skipped[f"unprovisioned:{o.module} {o.otype}"] += 1
            return True
        return False

    # helper: profile name -> profile object (for VS -> profile resolution)
    profiles: Dict[str, ConfigObject] = {}
    for o in objs:
        if o.module == "ltm" and o.otype.startswith("profile"):
            profiles[o.name] = o

    def add(fid: str, inst: str, note: str = "") -> None:
        f = feats[fid]
        if inst not in f.instances:
            f.instances.append(inst)
        if note and note not in f.detail:
            f.detail = (f.detail + "; " + note).strip("; ")

    for o in objs:
        k, n, b = o.kind, o.name, o.body
        if o.module == "ltm":
            comps.add(COMP_LTM)
            if o.otype == "virtual":
                add("virtual-server", n)
                if "source-address-translation" in b:
                    add("snat", n, "VS source-address-translation")
                for pname in re.findall(r"(/\S+)\s*\{[^}]*context clientside", b):
                    pass
                # profiles referenced
                for pref in re.findall(r"^\s{8}(/\S+)\s*\{", b, re.M):
                    p = profiles.get(pref)
                    if p:
                        _classify_profile(p, add, via=n)
                for r_ in re.findall(r"^\s{8}(/\S+)$", b, re.M):
                    pass
                if re.search(r"^\s+rules\s*\{", b, re.M):
                    for rn_ in re.findall(r"^\s{8}(/[\w/.-]+)$", b, re.M):
                        add("irule", f"{rn_} (on {n})")
                if "policies {" in b:
                    add("ltm-policy", n, "VS attaches LTM policy")
                if "persist {" in b:
                    pass  # persistence classified from profile objects
                if "security-log-profiles" in b:
                    add("security-logging", n)
                if "per-flow-request-access-policy" in b or "apm" in b:
                    pass
            elif o.otype == "pool":
                add("pool-lb", n)
            elif o.otype == "node":
                pass
            elif o.otype.startswith("monitor"):
                add("health-monitor", f"{n} ({o.otype.split()[-1]})")
            elif o.otype.startswith("profile"):
                _classify_profile(o, add)
            elif o.otype == "rule":
                add("irule", n)
                if "DNS::" in b or "DNS_REQUEST" in b:
                    add("dns-profile", n, "iRule uses DNS:: commands")
            elif o.otype == "policy":
                add("ltm-policy", n)
            elif o.otype.startswith("persistence"):
                ptype = o.otype.split()[-1]
                if ptype == "cookie":
                    add("persist-cookie", n,
                        "cookie-encryption required" if "cookie-encryption required" in b else "")
                elif ptype == "source-addr":
                    add("persist-srcaddr", n)
            elif o.otype in ("snatpool", "snat"):
                add("snat", n)
        elif o.module == "gtm":
            comps.add(COMP_GTM)
            if o.otype.startswith("wideip"):
                add("gtm-wideip", f"{n} ({o.otype.split()[-1].upper()})")
            elif o.otype.startswith("pool") or o.otype in ("server", "datacenter"):
                add("gtm-pool-server", f"{n} ({o.otype})")
            elif o.otype.startswith("monitor"):
                add("gtm-pool-server", f"{n} (monitor)")
            elif o.otype == "listener":
                add("gtm-listener", n)
                add("dns-profile", n, "listener DNS profile")
            elif "dnssec" in o.otype:
                add("dnssec", f"{n} ({o.otype})")
            elif o.otype in ("region", "topology", "global-settings load-balancing"):
                add("gtm-wideip", f"{n} ({o.otype})", "topology/region LB in use")
        elif o.module == "apm":
            if gated_out(o):
                continue
            comps.add(COMP_APM)
            add("apm", f"{n} ({o.otype})")
        elif o.module == "asm":
            if gated_out(o):
                continue
            comps.add(COMP_ASM)
            add("asm", n)
        elif o.module == "security":
            if gated_out(o):
                continue
            if "firewall" in o.otype:
                comps.add(COMP_AFM)
                add("afm", n)
            elif "bot-defense" in o.otype:
                comps.add(COMP_BOT)
                add("bot-defense", n)
            elif "log" in o.otype:
                add("security-logging", n)
        elif o.module == "sys":
            pass
    # dns profile objects
    for o in objs:
        if o.kind == "ltm profile dns":
            if "enable-dnssec yes" in o.body:
                add("dnssec", o.name, "ltm dns profile enable-dnssec")
            if "use-local-bind yes" in o.body:
                add("local-bind", o.name, "dns profile use-local-bind")
            if "enable-dns-express yes" in o.body:
                add("dns-profile", o.name, "DNS Express enabled")

    # management plane + HA always applicable on any BIG-IP
    add("mgmt-plane", "(system-wide: TMUI/tmsh/REST/mcpd)")
    add("upgrade-install", "(upgrade path 17.1.3 → target)")

    # provisioning as final authority over components discovered above
    if provisioned is not None:
        allowed = set(_BASELINE_COMPONENTS)
        for m in provisioned:
            allowed |= _PROVISION_COMPONENT_MAP.get(m, set())
        removed = comps - allowed
        comps &= allowed
        # drop features whose only components were unprovisioned
        for f in feats.values():
            if f.instances and not (f.components & comps):
                skipped[f"feature-dropped:{f.fid}"] += len(f.instances)
                f.instances = []
        if removed:
            print(f"  provisioning gate removed components: "
                  f"{', '.join(sorted(removed))}", file=sys.stderr)
    else:
        print("  WARNING: no 'sys provision' stanzas found — supply "
              "bigip_base.conf (--base-conf) for authoritative module gating; "
              "default-object denylist applied only", file=sys.stderr)

    active = [f for f in feats.values() if f.instances]
    prov_info: Dict[str, Any] = {"provision": prov, "skipped": dict(skipped)}
    return active, comps, prov_info


def _classify_profile(p: ConfigObject, add, via: str = "") -> None:
    suffix = f" (on {via})" if via else ""
    t = p.otype.replace("profile ", "")
    n = p.name + suffix
    if t == "client-ssl":
        add("client-ssl", n)
        if "ocsp-stapling enabled" in p.body:
            add("ocsp-stapling", n, "clientssl ocsp-stapling enabled")
        if "cert-key-chain" in p.body:
            add("cert-mgmt", n, "cert-key-chain configured")
    elif t == "server-ssl":
        add("server-ssl", n)
        if "peer-cert-mode require" in p.body:
            add("server-ssl", n)
    elif t == "http":
        add("http-profile", n)
    elif t == "http2":
        add("http2", n)
    elif t == "websocket":
        add("websocket", n)
    elif t == "one-connect":
        add("oneconnect", n)
    elif t == "tcp":
        add("tcp-profile", n)
    elif t == "udp":
        add("udp-profile", n)
    elif t == "dns":
        add("dns-profile", n)
    elif t == "access":
        add("apm", n)


# =============================================================================
# Release-notes HTML parser
# =============================================================================

@dataclass
class RNItem:
    item_id: str
    title: str = ""
    category: str = ""          # "Vulnerability Fixes" | "<X> Fixes" | "<X> Issues"
    known_issue: bool = False
    severity: str = ""
    cves: List[str] = field(default_factory=list)
    links: List[Tuple[str, str]] = field(default_factory=list)   # (label,url)
    description: str = ""
    fixed_versions: List[str] = field(default_factory=list)
    component: str = ""
    symptoms: str = ""
    conditions: str = ""
    impact: str = ""
    workaround: str = ""
    fix: str = ""


@dataclass
class ReleaseNotes:
    version: str
    build: str
    filename: str
    items: Dict[str, RNItem] = field(default_factory=dict)


_TAG = re.compile(r"<[^>]+>")


def _strip(s: str) -> str:
    return re.sub(r"\s+", " ", html_mod.unescape(_TAG.sub(" ", s))).strip()


_CAT_RE = re.compile(
    r'<font size="4"><u><strong>([^<]+)</strong></u></font>|'
    r'<font size="4"><b><u>([^<]+)</u></b></font>')
_ROW_RE = re.compile(r"<tr valign=\"top\">(.*?)</tr>", re.S)
_TD_RE = re.compile(r"<td[^>]*>(.*?)</td>", re.S)
_LINK_RE = re.compile(r'<a href="?([^">\s]+)"?[^>]*>([^<]+)</a>')
_DETAIL_RE = re.compile(
    r'<a name="A([\w.-]+)" rel="nofollow"></a>(.*?)(?=<a name="A[\w.-]+" rel="nofollow">|<a name="KnownIssues"|\Z)',
    re.S)
_FIELD_RE = re.compile(
    r"<strong>(Component|Symptoms|Conditions|Impact|Workaround|Fix|Fixed Versions|"
    r"Links to More Info):?\s*</strong>(.*?)(?=<p><strong>|</div>|\Z)", re.S)


def parse_release_notes(path: str) -> ReleaseNotes:
    data = open(path, encoding="utf-8", errors="replace").read()
    ver = re.search(r"Version:\s*</b>?\s*([\d.]+)", _strip(data[:4000]) ) or \
          re.search(r"Version:\s*([\d.]+)", _strip(data[:4000]))
    build = re.search(r"Build:\s*([\d.]+)", _strip(data[:4000]))
    rn = ReleaseNotes(version=ver.group(1) if ver else "?",
                      build=build.group(1) if build else "?",
                      filename=os.path.basename(path))

    # ---- pass 1: summary tables, walked sequentially with category context
    pos = 0
    cats = [(m.start(), (m.group(1) or m.group(2)).strip())
            for m in _CAT_RE.finditer(data)]
    for idx, (start, cat) in enumerate(cats):
        end = cats[idx + 1][0] if idx + 1 < len(cats) else len(data)
        seg = data[start:end]
        if "Known Issues in BIG-IP" in cat:
            continue
        is_issue = cat.endswith("Issues")
        for row in _ROW_RE.finditer(seg):
            tds = _TD_RE.findall(row.group(1))
            if len(tds) < 3:
                continue
            first = _strip(tds[0])
            if first in ("ID Number", ""):
                continue
            item_id = first
            it = rn.items.get(item_id) or RNItem(item_id=item_id)
            it.category = cat
            it.known_issue = is_issue
            col2 = _strip(tds[1])
            if col2.startswith("CVE-") or ("Vulnerability" in cat and "CVE" in col2):
                it.cves = re.findall(r"CVE-\d{4}-\d+", col2)
            else:
                it.severity = col2
            if len(tds) >= 4:
                it.links = [(lbl, url) for url, lbl in _LINK_RE.findall(tds[2])]
                it.description = _strip(tds[3])
            if not is_issue and len(tds) >= 5:
                it.fixed_versions = [v.strip() for v in _strip(tds[4]).split(",")
                                     if re.match(r"^\d", v.strip())]
            rn.items[item_id] = it

    # ---- pass 2: detail blocks (Component / Symptoms / Conditions / ...)
    for m in _DETAIL_RE.finditer(data):
        item_id, block = m.group(1), m.group(2)
        it = rn.items.get(item_id) or RNItem(item_id=item_id)
        tm = re.search(r"</a>\s*" + re.escape(item_id) + r"\s*:\s*(.*?)</h4>",
                       m.group(0), re.S)
        if tm:
            it.title = _strip(tm.group(1))
        for fm in _FIELD_RE.finditer(block):
            fname, fval = fm.group(1), _strip(fm.group(2))
            if fname == "Component":
                it.component = fval
            elif fname == "Symptoms":
                it.symptoms = fval
            elif fname == "Conditions":
                it.conditions = fval
            elif fname == "Impact":
                it.impact = fval
            elif fname == "Workaround":
                it.workaround = fval
            elif fname == "Fix":
                it.fix = fval
            elif fname == "Fixed Versions" and not it.fixed_versions:
                it.fixed_versions = [v.strip() for v in fval.split(",")
                                     if re.match(r"^\d", v.strip())]
        if not it.cves:
            it.cves = sorted(set(re.findall(r"CVE-\d{4}-\d+",
                                            it.title + " " + it.description)))
        rn.items[item_id] = it
    return rn


# =============================================================================
# Version logic
# =============================================================================

def _vtuple(v: str) -> Tuple[int, ...]:
    return tuple(int(x) for x in re.findall(r"\d+", v)[:4])


def already_remediated(item: RNItem, current: str) -> bool:
    """True if the item is fixed at or below the current running version
    on the same maintenance branch (e.g. fixed in 17.1.3 while running 17.1.3;
    fixed only in 17.1.3.1 => NOT remediated)."""
    cur = _vtuple(current)
    branch = cur[:2]
    for v in item.fixed_versions:
        t = _vtuple(v)
        if t[:2] == branch:
            a = t + (0,) * (4 - len(t))
            b = cur + (0,) * (4 - len(cur))
            if a <= b:
                return True
    return False


# =============================================================================
# Platform-wide (config-independent) CVE classifier
# =============================================================================

_PLATFORM_PKGS = re.compile(
    r"\b(kernel|linux kernel|postgresql|postgres|apache http|apache vulnerab|httpd|"
    r"openssh|libssh|ssh\b|python|urllib3|lodash|node\.?js|jquery|grub2?|libxml2|"
    r"expat|glibc|systemd|sudo|curl\b|libcurl|samba|vim\b|bash\b|rsync|zlib|"
    r"xz\b|krb5|kerberos library|nss\b|gnutls|binutils|ncurses|sqlite|qt\b|gson|"
    r"java\b|openjdk|phantomjs|iputils|centos|host\s?os|hostos|f5os)\b", re.I)
_BIND_RE = re.compile(r"\bbind\b|\bnamed\b", re.I)


def is_platform_wide(item: RNItem, dns_in_use: bool) -> bool:
    if not item.cves:
        return False
    txt = " ".join([item.title, item.description, item.symptoms])
    if _BIND_RE.search(txt):
        return not dns_in_use      # BIND CVEs are config-relevant when DNS/GTM used
    return bool(_PLATFORM_PKGS.search(txt))


# =============================================================================
# Matching engine (deterministic)
# =============================================================================

@dataclass
class Match:
    feature: Feature
    item: RNItem
    release: str                 # target release version
    score: int
    band: str                    # exact / probable / component / weak
    reason: str


_SEV_COMPONENT_SCORE = {
    "1-Blocking": 55, "2-Critical": 50, "3-Major": 42, "4-Minor": 35,
}


def _compile(terms: List[str]) -> List[re.Pattern]:
    return [re.compile(t, re.I) for t in terms]


def score_item(feat: Feature, item: RNItem,
               strong_re: List[re.Pattern], weak_re: List[re.Pattern],
               comps_in_config: Set[str]) -> Tuple[int, str, str]:
    cond = item.conditions if item.conditions.lower() not in ("na", "n/a", "none", "") \
        else ""
    symtitle = " ".join([item.title, item.symptoms, item.description])

    s_cond = sorted({p.pattern for p in strong_re if cond and p.search(cond)})
    s_sym = sorted({p.pattern for p in strong_re if p.search(symtitle)})
    w_all = sorted({p.pattern for p in weak_re
                    if p.search(cond) or p.search(symtitle)})

    if s_cond:
        score = min(100, 85 + 5 * (len(s_cond) - 1) + 2 * min(len(w_all), 3)
                    + (3 if s_sym else 0))
        return score, "exact", (f"strong term(s) {_fmt_terms(s_cond)} in Conditions"
                                + (f"; also in Symptoms/Title" if s_sym else ""))
    if s_sym:
        score = min(84, 60 + 6 * (len(s_sym) - 1) + 2 * min(len(w_all), 4))
        return score, "probable", f"strong term(s) {_fmt_terms(s_sym)} in Symptoms/Title"
    comp_hit = item.component in feat.components and item.component in comps_in_config
    if comp_hit:
        base = _SEV_COMPONENT_SCORE.get(item.severity, 32)
        score = min(59, base + 2 * min(len(w_all), 3))
        return score, "component", (f"component match ({item.component}); "
                                    f"no feature-specific terms")
    if w_all:
        return min(29, 5 + 6 * len(w_all)), "weak", \
            f"weak lexical hit(s) {_fmt_terms(w_all)} only"
    return 0, "none", ""


def _fmt_terms(pats: List[str]) -> str:
    clean = [re.sub(r"\\b|\(\?[a-z]+\)|[\\^$]", "", p).replace(".?", " ")
             for p in pats[:4]]
    return "'" + "', '".join(clean) + "'"


def run_matching(features: List[Feature], releases: List[ReleaseNotes],
                 comps: Set[str], current_version: str
                 ) -> Tuple[List[Match], List[Tuple[str, RNItem]],
                            List[Tuple[str, RNItem]], Dict[str, List[RNItem]]]:
    """Returns (matches, platform_cves, remediated, unmatched_items_per_release)."""
    dns_in_use = COMP_GTM in comps or any(f.fid.startswith(("dns", "gtm", "local-bind"))
                                          for f in features)
    compiled = {f.fid: (_compile(f.strong), _compile(f.weak)) for f in features}

    matches: List[Match] = []
    platform: List[Tuple[str, RNItem]] = []
    remediated: List[Tuple[str, RNItem]] = []
    unmatched: Dict[str, List[RNItem]] = defaultdict(list)
    seen_platform: Set[str] = set()
    seen_remediated: Set[str] = set()

    for rel in releases:
        for item in rel.items.values():
            if not item.known_issue and already_remediated(item, current_version):
                if item.item_id not in seen_remediated:
                    remediated.append((rel.version, item))
                    seen_remediated.add(item.item_id)
                continue
            if not item.known_issue and is_platform_wide(item, dns_in_use):
                if item.item_id not in seen_platform:
                    platform.append((rel.version, item))
                    seen_platform.add(item.item_id)
                continue
            scored: List[Match] = []
            for f in features:
                sre, wre = compiled[f.fid]
                score, band, reason = score_item(f, item, sre, wre, comps)
                if score > 0:
                    scored.append(Match(f, item, rel.version, score, band, reason))
            if not scored:
                unmatched[rel.version].append(item)
                continue
            strong = sorted([m for m in scored if m.band in ("exact", "probable")],
                            key=lambda m: -m.score)[:3]
            if strong:
                matches.extend(strong)
            else:
                comp_m = [m for m in scored if m.band == "component"]
                weak_m = [m for m in scored if m.band == "weak"]
                if comp_m:
                    # one row per item: the component itself is the association,
                    # attach the config feature with the most instances
                    best = max(comp_m, key=lambda m: (m.score, len(m.feature.instances)))
                    matches.append(best)
                elif weak_m:
                    matches.append(max(weak_m, key=lambda m: m.score))
    # merge identical (item, feature) matches across target releases
    merged: Dict[Tuple[str, str], Match] = OrderedDict()
    for m in matches:
        key = (m.item.item_id, m.feature.fid)
        if key in merged:
            prev = merged[key]
            rels_ = prev.release.split(", ")
            if m.release not in rels_:
                rels_.append(m.release)
                prev.release = ", ".join(sorted(rels_, key=_vtuple))
            if m.score > prev.score:
                prev.score, prev.band, prev.reason = m.score, m.band, m.reason
        else:
            merged[key] = m
    matches = list(merged.values())
    matches.sort(key=lambda m: (-m.score, m.item.item_id))
    return matches, platform, remediated, unmatched


# =============================================================================
# Release comparison / recommendation
# =============================================================================

def _fixed_in_target(item: RNItem, target_rel: "ReleaseNotes") -> bool:
    """Fixed when moving to target_rel: item appears in the target's own
    cumulative fix list, or fixed_versions contains a version on the target's
    major.minor branch at or below the target."""
    if item.item_id in target_rel.items and not target_rel.items[item.item_id].known_issue:
        return True
    t = _vtuple(target_rel.version)
    for v in item.fixed_versions:
        vt = _vtuple(v)
        if vt[:2] == t[:2] and (vt + (0,) * 4)[:4] <= (t + (0,) * 4)[:4]:
            return True
    return False


def compare_releases(matches: List["Match"], platform: List[Tuple[str, "RNItem"]],
                     releases: List["ReleaseNotes"]) -> List[Dict[str, Any]]:
    """Per target release: strong config-matched bugs fixed, residual (matched
    elsewhere but NOT fixed by this target), known-issue upgrade risks, and
    platform CVEs resolved. Risk score = residual + known-issue risks.
    Lowest risk score => recommended (tie-break: most bugs fixed)."""
    strong_fix: Dict[str, RNItem] = {}
    for m in matches:
        if not m.item.known_issue and m.score >= 60:
            strong_fix[m.item.item_id] = m.item
    strong_risk: Dict[str, Tuple[RNItem, Set[str]]] = {}
    for m in matches:
        if m.item.known_issue and m.score >= 60:
            rels_ = set(m.release.split(", "))
            if m.item.item_id in strong_risk:
                strong_risk[m.item.item_id][1].update(rels_)
            else:
                strong_risk[m.item.item_id] = (m.item, rels_)

    rows: List[Dict[str, Any]] = []
    for rel in releases:
        fixed = [it for it in strong_fix.values() if _fixed_in_target(it, rel)]
        residual = [it for it in strong_fix.values() if not _fixed_in_target(it, rel)]
        risks = [it for it, rs in strong_risk.values() if rel.version in rs]
        plat_fixed = [it for _, it in platform if _fixed_in_target(it, rel)]
        rows.append({
            "version": rel.version,
            "fixed": len(fixed), "residual": len(residual),
            "risks": len(risks), "platform_fixed": len(plat_fixed),
            "risk_score": len(residual) + len(risks),
        })
    best = min(rows, key=lambda r: (r["risk_score"], -r["fixed"]))
    for r in rows:
        r["recommended"] = r is best
    return rows


# =============================================================================
# Susceptibility heat map
# =============================================================================
# Susceptibility of a config feature in a target release = weighted sum of
# matched RN items that remain PRESENT after upgrading to that release:
#   - known issues applicable to the target branch, plus
#   - fix-items matched to the config but NOT fixed in that target (residual).
# Band weights: exact 3.0, probable 2.0, component 1.0, weak 0.25.

_HEAT_WEIGHTS = {"exact": 3.0, "probable": 2.0, "component": 1.0, "weak": 0.25}


def compute_heatmap(matches: List["Match"], releases: List["ReleaseNotes"]
                    ) -> Tuple[List[Dict[str, Any]], float]:
    rel_by_ver = {r.version: r for r in releases}
    rows: Dict[str, Dict[str, Any]] = OrderedDict()
    for m in matches:
        row = rows.setdefault(m.feature.fid, {
            "fid": m.feature.fid, "label": m.feature.label,
            "instances": len(m.feature.instances),
            "cells": {r.version: {"index": 0.0, "exact": 0, "probable": 0,
                                  "component": 0, "weak": 0}
                      for r in releases},
        })
        m_rels = m.release.split(", ")
        for rel in releases:
            rv = rel.version
            if m.item.known_issue:
                present = rv in m_rels
            else:
                present = not _fixed_in_target(m.item, rel_by_ver[rv])
            if present:
                c = row["cells"][rv]
                c[m.band] += 1
                c["index"] += _HEAT_WEIGHTS[m.band]
    out = list(rows.values())
    for r in out:
        r["total"] = sum(c["index"] for c in r["cells"].values())
    out.sort(key=lambda r: -r["total"])
    vmax = max((c["index"] for r in out for c in r["cells"].values()),
               default=1.0) or 1.0
    return out, vmax


def _heat_color(v: float, vmax: float) -> Tuple[str, str]:
    """Continuous green -> yellow -> red ramp; returns (bg_hex, fg_hex)."""
    t = max(0.0, min(1.0, v / vmax))
    stops = [(0.0, (0x1e, 0x8e, 0x3e)), (0.5, (0xf9, 0xab, 0x00)),
             (1.0, (0xd9, 0x30, 0x25))]
    for (t0, c0), (t1, c1) in zip(stops, stops[1:]):
        if t <= t1:
            f = 0.0 if t1 == t0 else (t - t0) / (t1 - t0)
            rgb = tuple(round(a + (b - a) * f) for a, b in zip(c0, c1))
            break
    else:
        rgb = stops[-1][1]
    lum = 0.299 * rgb[0] + 0.587 * rgb[1] + 0.114 * rgb[2]
    fg = "#202124" if lum > 150 else "#ffffff"
    return "#%02x%02x%02x" % rgb, fg


# =============================================================================
# Report generation — HTML
# =============================================================================

def _color(score: int) -> Tuple[str, str]:
    """(bg, fg) for a score chip. Severity-oriented: stronger match = more
    alarming color (exact=red, probable=dark orange, component=yellow,
    weak=light green)."""
    if score >= 85:
        return "#d93025", "#ffffff"      # red — exact match, highest concern
    if score >= 60:
        return "#e8710a", "#ffffff"      # dark orange — probable
    if score >= 30:
        return "#f9ab00", "#202124"      # yellow — component-level
    if score >= 1:
        return "#81c995", "#202124"      # light green — weak match
    return "#9aa0a6", "#ffffff"          # gray — no association


def _chip(score: int) -> str:
    bg, fg = _color(score)
    return (f'<span class="chip" style="background:{bg};color:{fg}">'
            f'{score}</span>')


def _esc(s: str) -> str:
    return html_mod.escape(s or "", quote=True)


def _links_html(item: RNItem) -> str:
    out = []
    for lbl, url in item.links:
        if url.startswith(("http://", "https://")):
            out.append(f'<a href="{_esc(url)}" target="_blank" rel="noopener">'
                       f'{_esc(lbl)}</a>')
    return ", ".join(out) or "&mdash;"


def build_html(features: List[Feature], releases: List[ReleaseNotes],
               matches: List[Match], platform: List[Tuple[str, RNItem]],
               remediated: List[Tuple[str, RNItem]],
               unmatched: Dict[str, List[RNItem]],
               comps: Set[str], current_version: str, hostname: str,
               comparison: Optional[List[Dict[str, Any]]] = None) -> str:
    comparison = comparison or []
    rel_versions = [r.version for r in releases]
    gen = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")

    fix_matches = [m for m in matches if not m.item.known_issue]
    risk_matches = [m for m in matches if m.item.known_issue]

    # summary matrix: feature x release -> best score + counts per band
    def band_counts(ms: List[Match]) -> Dict[Tuple[str, str], Dict[str, int]]:
        d: Dict[Tuple[str, str], Dict[str, int]] = defaultdict(
            lambda: {"exact": 0, "probable": 0, "component": 0, "weak": 0, "best": 0})
        for m in ms:
            for rv in m.release.split(", "):
                c = d[(m.feature.fid, rv)]
                c[m.band] += 1
                c["best"] = max(c["best"], m.score)
        return d
    matrix = band_counts(fix_matches)

    css = """
    body{font-family:'Segoe UI',Roboto,Helvetica,Arial,sans-serif;margin:0;
         background:#f5f6f8;color:#202124}
    header{background:#0b2e4f;color:#fff;padding:18px 28px}
    header h1{margin:0;font-size:20px} header .sub{color:#bcd2e8;font-size:13px;margin-top:4px}
    nav{position:sticky;top:0;background:#fff;border-bottom:1px solid #dadce0;
        padding:8px 28px;z-index:5;font-size:13px}
    nav a{margin-right:16px;color:#0b57d0;text-decoration:none}
    main{padding:20px 28px;max-width:1500px}
    h2{font-size:16px;border-bottom:2px solid #0b2e4f;padding-bottom:4px;margin-top:34px}
    table{border-collapse:collapse;width:100%;background:#fff;font-size:12.5px;
          box-shadow:0 1px 2px rgba(0,0,0,.08)}
    th{background:#eef1f5;text-align:left;padding:6px 8px;border-bottom:2px solid #dadce0;
       position:sticky;top:37px}
    td{padding:6px 8px;border-bottom:1px solid #eceff3;vertical-align:top}
    tr:hover td{background:#f8fbff}
    .chip{display:inline-block;min-width:34px;text-align:center;border-radius:12px;
          padding:2px 8px;font-weight:700;font-size:12px}
    .band-exact{border-left:4px solid #d93025}.band-probable{border-left:4px solid #e8710a}
    .band-component{border-left:4px solid #f9ab00}.band-weak{border-left:4px solid #81c995}
    .mono{font-family:Consolas,Menlo,monospace;font-size:12px}
    .tag{display:inline-block;background:#e8f0fe;color:#174ea6;border-radius:4px;
         padding:1px 6px;margin:1px 2px;font-size:11px}
    .cve{background:#fce8e6;color:#a50e0e}
    .ki{background:#fef7e0;color:#7a5c00}
    details{margin:6px 0} details summary{cursor:pointer;font-weight:600;font-size:13px}
    .legend span{margin-right:14px}
    .small{color:#5f6368;font-size:11.5px}
    .kpi{display:inline-block;background:#fff;border:1px solid #dadce0;border-radius:8px;
         padding:10px 16px;margin:6px 10px 6px 0;text-align:center}
    .kpi b{display:block;font-size:22px}
    footer{padding:18px 28px;color:#5f6368;font-size:11px}
    """

    def item_rows(ms: List[Match]) -> str:
        rows = []
        for m in ms:
            it = m.item
            cve = " ".join(f'<span class="tag cve">{_esc(c)}</span>' for c in it.cves)
            ki = '<span class="tag ki">KNOWN ISSUE</span>' if it.known_issue else ""
            inst = ", ".join(_esc(i) for i in m.feature.instances[:6])
            more = (f" <span class='small'>(+{len(m.feature.instances)-6} more)</span>"
                    if len(m.feature.instances) > 6 else "")
            fv = ", ".join(it.fixed_versions) or "&mdash;"
            rows.append(
                f'<tr class="band-{m.band}">'
                f"<td>{_chip(m.score)}</td>"
                f"<td class='mono'>{_esc(it.item_id)}</td>"
                f"<td>{_esc(m.release)}</td>"
                f"<td>{_esc(m.feature.label)}<div class='small'>{inst}{more}</div></td>"
                f"<td>{_esc(it.title or it.description)} {cve} {ki}"
                f"<div class='small'>{_esc(it.component or m.item.category)}"
                f" &middot; {_esc(it.severity) if it.severity else 'CVE'}</div></td>"
                f"<td>{_esc(m.reason)}</td>"
                f"<td>{_links_html(it)}</td>"
                f"<td class='small'>{fv}</td></tr>")
        return "".join(rows)

    hdr = ("<tr><th>Score</th><th>Bug ID</th><th>Target Rel.</th>"
           "<th>Config Element (feature &middot; instances)</th>"
           "<th>Release-Note Item</th><th>Match Rationale</th>"
           "<th>Links</th><th>Fixed In</th></tr>")

    strong_fix = [m for m in fix_matches if m.band in ("exact", "probable")]
    comp_fix = [m for m in fix_matches if m.band == "component"]
    weak_fix = [m for m in fix_matches if m.band == "weak"]

    # matrix table
    mat_rows = []
    for f in features:
        cells = []
        for rv in rel_versions:
            c = matrix.get((f.fid, rv))
            if not c:
                cells.append("<td class='small'>&mdash;</td>")
            else:
                parts = [f"{c[b]} {lbl}" for b, lbl in
                         (("exact", "exact"), ("probable", "probable"),
                          ("component", "component-level"), ("weak", "weak"))
                         if c[b]]
                cells.append(f"<td>{_chip(c['best'])} "
                             f"<span class='small'>best score &middot; "
                             f"{', '.join(parts)}</span></td>")
        mat_rows.append(f"<tr><td>{_esc(f.label)}"
                        f"<div class='small'>{len(f.instances)} instance(s)</div></td>"
                        + "".join(cells) + "</tr>")

    platform_rows = "".join(
        f"<tr><td class='mono'>{_esc(it.item_id)}</td><td>{_esc(rv)}</td>"
        f"<td>{' '.join(f'<span class=\"tag cve\">{_esc(c)}</span>' for c in it.cves)}</td>"
        f"<td>{_esc(it.title or it.description)}</td><td>{_links_html(it)}</td>"
        f"<td class='small'>{', '.join(it.fixed_versions) or '&mdash;'}</td></tr>"
        for rv, it in sorted(platform, key=lambda x: x[1].item_id))

    remed_rows = "".join(
        f"<tr><td class='mono'>{_esc(it.item_id)}</td>"
        f"<td>{_esc(it.title or it.description)}</td>"
        f"<td class='small'>{', '.join(it.fixed_versions)}</td></tr>"
        for rv, it in sorted(remediated, key=lambda x: x[1].item_id))

    unmatched_html = []
    for rv in rel_versions:
        us = unmatched.get(rv, [])
        rows = "".join(
            f"<tr><td class='mono'>{_esc(it.item_id)}</td>"
            f"<td>{_esc(it.component or it.category)}</td>"
            f"<td>{_esc(it.severity or ('CVE' if it.cves else ''))}</td>"
            f"<td>{_esc(it.title or it.description)}"
            f"{' <span class=\"tag ki\">KNOWN ISSUE</span>' if it.known_issue else ''}"
            f"</td><td>{_links_html(it)}</td></tr>" for it in us)
        unmatched_html.append(
            f"<details><summary>Target {_esc(rv)} — {len(us)} items with no "
            f"association to this configuration</summary>"
            f"<table><tr><th>Bug ID</th><th>Component</th><th>Sev</th>"
            f"<th>Item</th><th>Links</th></tr>{rows}</table></details>")

    heat_rows, heat_max = compute_heatmap(matches, releases)
    hm = []
    for r in heat_rows:
        cells = []
        for rv in rel_versions:
            c = r["cells"][rv]
            bg, fg = _heat_color(c["index"], heat_max)
            parts = [f"{c[b]} {lbl}" for b, lbl in
                     (("exact", "exact"), ("probable", "probable"),
                      ("component", "comp"), ("weak", "weak")) if c[b]]
            cells.append(
                f"<td style='background:{bg};color:{fg};text-align:center' "
                f"title='{_esc(', '.join(parts) or 'no residual exposure')}'>"
                f"<b>{c['index']:g}</b><br>"
                f"<span style='font-size:10px'>{_esc(', '.join(parts) or '&mdash;')}"
                f"</span></td>")
        hm.append(f"<tr><td>{_esc(r['label'])}"
                  f"<div class='small'>{r['instances']} instance(s)</div></td>"
                  + "".join(cells) + "</tr>")
    heatmap_html = ("<table><tr><th>Config Feature (most &rarr; least "
                    "susceptible)</th>"
                    + "".join(f"<th style='text-align:center'>{_esc(v)}</th>"
                              for v in rel_versions)
                    + "</tr>" + "".join(hm) + "</table>")

    comp_rows_html = []
    for r in comparison:
        star = " &#9733; RECOMMENDED" if r["recommended"] else ""
        style = "background:#e6f4ea;font-weight:700" if r["recommended"] else ""
        comp_rows_html.append(
            f"<tr style='{style}'><td>{_esc(r['version'])}{star}</td>"
            f"<td>{r['fixed']}</td><td>{r['residual']}</td><td>{r['risks']}</td>"
            f"<td>{r['platform_fixed']}</td><td><b>{r['risk_score']}</b></td></tr>")
    comparison_html = (
        "<table><tr><th>Target Release</th>"
        "<th>Config-matched bugs FIXED by upgrading</th>"
        "<th>Residual: matched bugs NOT fixed in this release</th>"
        "<th>Known-issue upgrade risks matched to config</th>"
        "<th>Platform CVEs resolved</th>"
        "<th>Risk score (residual + known-issue risks) — lower is safer</th></tr>"
        + "".join(comp_rows_html) + "</table>"
        "<p class='small'>Counts use strong associations only (score &ge; 60). "
        "'Residual' = a defect that matches this configuration and is fixed in "
        "another target release but <b>not</b> in this one — choosing this "
        "release leaves that bug in place. Recommendation = lowest risk score, "
        "tie-broken by most bugs fixed. <b>Caveat:</b> newer branches "
        "typically publish shorter known-issue lists partly due to less field "
        "exposure — weigh maturity alongside these counts.</p>")

    def per_release_tables(ms: List[Match], open_first: bool = True) -> str:
        out = []
        for i, rv in enumerate(rel_versions):
            sub = [m for m in ms if rv in m.release.split(", ")]
            op = " open" if (open_first and i == 0) else ""
            out.append(f"<details{op}><summary>Target {_esc(rv)} — "
                       f"{len(sub)} matches</summary>"
                       f"<table>{hdr}{item_rows(sub)}</table></details>")
        return "".join(out)

    feat_tags = "".join(f"<span class='tag'>{_esc(f.label)} "
                        f"({len(f.instances)})</span>" for f in features)
    comp_tags = "".join(f"<span class='tag'>{_esc(c)}</span>" for c in sorted(comps))

    html_out = f"""<!DOCTYPE html><html lang="en"><head><meta charset="utf-8">
<title>TMOS Release-Note Advisor — {_esc(hostname)} ({_esc(current_version)})</title>
<style>{css}</style></head><body>
<header><h1>TMOS Release-Note Configuration Advisor</h1>
<div class="sub">Device: <b>{_esc(hostname)}</b> &middot; running TMOS
<b>{_esc(current_version)}</b> &middot; target releases:
<b>{_esc(', '.join(rel_versions))}</b> &middot; generated {gen} &middot;
engine v{__version__} (deterministic, offline)</div></header>
<nav><a href="#summary">Summary</a><a href="#compare">Comparison</a><a href="#heatmap">Heat Map</a><a href="#matrix">Feature Matrix</a>
<a href="#assoc">Associations</a><a href="#compband">Component-Level</a>
<a href="#risks">Upgrade Risks</a><a href="#platform">Platform CVEs</a>
<a href="#remediated">Already Remediated</a><a href="#unmatched">Unmatched</a>
<a href="#method">Methodology</a></nav>
<main>
<h2 id="summary">Summary</h2>
<div>
<div class="kpi"><b>{len(features)}</b>config features detected</div>
<div class="kpi"><b>{sum(len(f.instances) for f in features)}</b>named instances</div>
<div class="kpi"><b>{sum(len(r.items) for r in releases)}</b>release-note items parsed</div>
<div class="kpi"><b>{len(strong_fix)}</b>strong associations (&ge;60)</div>
<div class="kpi"><b>{len(risk_matches)}</b>upgrade-risk matches (known issues)</div>
<div class="kpi"><b>{len(platform)}</b>platform-wide CVEs</div>
<div class="kpi"><b>{len(remediated)}</b>already remediated on {_esc(current_version)}</div>
</div>
<p class="legend"><b>Score bands:</b>
<span>{_chip(92)} 85–100 exact — feature term in <i>Conditions</i></span>
<span>{_chip(72)} 60–84 probable — feature term in Symptoms/Title</span>
<span>{_chip(45)} 30–59 component-level only</span>
<span>{_chip(15)} 1–29 weak lexical</span></p>
<p><b>Detected features:</b> {feat_tags}</p>
<p><b>Modules/components in configuration</b>
<span class="small">(gated by <code>sys provision</code>; default/system
objects excluded)</span><b>:</b> {comp_tags}</p>

<h2 id="compare">Target-Release Comparison — Which Version Is Safest?</h2>
{comparison_html}

<h2 id="heatmap">Configuration Susceptibility Heat Map — Post-Upgrade Bug
Exposure per Target Release</h2>
<p class="small">Cell value = susceptibility index: weighted sum of matched
release-note items that <b>remain present after upgrading</b> to that release
(known issues on the target branch + config-matched bugs not fixed by the
target). Weights: exact &times;3, probable &times;2, component &times;1,
weak &times;0.25. Ordered most &rarr; least susceptible; color normalized
across the matrix (green = least, red = most).</p>
{heatmap_html}

<h2 id="matrix">Feature × Target-Release Matrix (best score; E=exact P=probable C=component W=weak)</h2>
<table><tr><th>Config Feature</th>{''.join(f'<th>{_esc(v)}</th>' for v in rel_versions)}</tr>
{''.join(mat_rows)}</table>

<h2 id="assoc">Direct Associations — Fixes &amp; Vulnerabilities (score &ge; 60)</h2>
<p class="small">Broken out per target release; a bug present in multiple
cumulative release-note files appears under each applicable release.</p>
{per_release_tables(strong_fix)}
<details><summary>All target releases — consolidated ({len(strong_fix)} rows)</summary>
<table>{hdr}{item_rows(strong_fix)}</table></details>

<h2 id="compband">Component-Level Associations (30–59)</h2>
<p class="small">The configuration uses the item's component, but no
feature-specific term matched. Review when planning the upgrade; individually
lower confidence by design.</p>
<details><summary>{len(comp_fix)} component-level matches</summary>
<table>{hdr}{item_rows(comp_fix)}</table></details>
<details><summary>{len(weak_fix)} weak lexical matches (1–29)</summary>
<table>{hdr}{item_rows(weak_fix)}</table></details>

<h2 id="risks">Upgrade Risks — Known Issues in Target Releases Matched to This Config</h2>
<p class="small">These defects exist (unfixed) in the target release train and
your configuration exercises the associated feature. Weigh these before upgrading.</p>
{per_release_tables(risk_matches)}
<details><summary>All target releases — consolidated ({len(risk_matches)} rows)</summary>
<table>{hdr}{item_rows(risk_matches)}</table></details>

<h2 id="platform">Platform-Wide CVEs (config-independent — apply to every
{_esc(current_version)} device)</h2>
<p class="small">Host-OS / bundled-package vulnerabilities (kernel, PostgreSQL,
Apache, OpenSSH, Python, …). Not scored against configuration by design;
all are resolved by moving to a release listed in "Fixed In".</p>
<table><tr><th>Bug ID</th><th>Target Rel.</th><th>CVE</th><th>Item</th>
<th>Links</th><th>Fixed In</th></tr>{platform_rows}</table>

<h2 id="remediated">Already Remediated on {_esc(current_version)}</h2>
<p class="small">Items whose fixed versions include {_esc(current_version)} or
lower on the {'.'.join(current_version.split('.')[:2])}.x branch — no action needed.</p>
<details><summary>{len(remediated)} items</summary>
<table><tr><th>Bug ID</th><th>Item</th><th>Fixed In</th></tr>{remed_rows}</table></details>

<h2 id="unmatched">No Association to This Configuration</h2>
{''.join(unmatched_html)}

<h2 id="method">Methodology</h2>
<p class="small">Deterministic engine: tmsh configuration parsed brace-aware into
objects; features detected via taxonomy (profiles, persistence, iRule events,
SSL options, GTM/DNSSEC objects, module presence). Release-note HTML parsed into
per-item records (category, CVE/severity, component, symptoms, <b>conditions</b>,
impact, workaround, fix, fixed versions). Scoring weights the Conditions field
highest because it states the configuration required to trigger the defect.
BIND/named CVEs are treated as configuration-relevant when DNS/GTM is in use,
platform-wide otherwise. No data leaves this host; no LLM/semantic inference.</p>
</main>
<footer>tmos_rn_advisor v{__version__} — read-only analysis. Verify all upgrade
decisions against F5 K-articles linked above. Not affiliated with F5, Inc.</footer>
</body></html>"""
    return html_out


# =============================================================================
# Report generation — Excel
# =============================================================================

def build_xlsx(path: str, features: List[Feature], releases: List[ReleaseNotes],
               matches: List[Match], platform: List[Tuple[str, RNItem]],
               remediated: List[Tuple[str, RNItem]],
               unmatched: Dict[str, List[RNItem]],
               current_version: str, hostname: str,
               comparison: Optional[List[Dict[str, Any]]] = None) -> None:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = Workbook()

    def fill_for(score: int) -> PatternFill:
        bg, _ = _color(score)
        return PatternFill("solid", fgColor=bg.lstrip("#").upper())

    hdr_font = Font(bold=True, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="0B2E4F")

    def sheet(ws, headers: List[str], widths: List[int]):
        ws.append(headers)
        for c in range(1, len(headers) + 1):
            cell = ws.cell(row=1, column=c)
            cell.font, cell.fill = hdr_font, hdr_fill
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1]
        ws.freeze_panes = "A2"

    # Summary
    ws = wb.active; ws.title = "Summary"
    ws.append(["TMOS Release-Note Configuration Advisor"])
    ws["A1"].font = Font(bold=True, size=14)
    for row in [
        ["Device", hostname], ["Current TMOS", current_version],
        ["Target releases", ", ".join(r.version for r in releases)],
        ["Generated (UTC)", datetime.now(timezone.utc).isoformat(timespec="seconds")],
        ["Engine", f"tmos_rn_advisor v{__version__} (deterministic)"],
        [], ["Config features detected", len(features)],
        ["Named instances", sum(len(f.instances) for f in features)],
        ["RN items parsed", sum(len(r.items) for r in releases)],
        ["Strong associations (>=60)", sum(1 for m in matches
                                           if not m.item.known_issue and m.score >= 60)],
        ["Upgrade-risk matches", sum(1 for m in matches if m.item.known_issue)],
        ["Platform-wide CVEs", len(platform)],
        [f"Already remediated on {current_version}", len(remediated)],
    ]:
        ws.append(row)

    # --- Target-release comparison / recommendation ---
    ws.append([])
    ws.append(["TARGET-RELEASE COMPARISON — lower risk score = safer upgrade"])
    ws.cell(row=ws.max_row, column=1).font = Font(bold=True, size=12)
    comp_hdr_row = ws.max_row + 1
    ws.append(["Target Release", "Config-matched bugs FIXED",
               "Residual bugs NOT fixed in this release",
               "Known-issue upgrade risks", "Platform CVEs resolved",
               "RISK SCORE (residual + risks)", "Verdict"])
    for c in range(1, 8):
        cell = ws.cell(row=comp_hdr_row, column=c)
        cell.font, cell.fill = hdr_font, hdr_fill
    rec_fill = PatternFill("solid", fgColor="C6EFCE")
    for r in (comparison or []):
        ws.append([r["version"], r["fixed"], r["residual"], r["risks"],
                   r["platform_fixed"], r["risk_score"],
                   "RECOMMENDED (safest)" if r["recommended"] else ""])
        if r["recommended"]:
            for c in range(1, 8):
                ws.cell(row=ws.max_row, column=c).fill = rec_fill
                ws.cell(row=ws.max_row, column=c).font = Font(bold=True)
    ws.append(["Residual = a defect matching this configuration that is fixed in"
               " another target release but NOT in this one (strong matches"
               " >=60 only)."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9)

    # --- Score-band legend (row numbers captured dynamically) ---
    ws.append([])
    ws.append(["Score bands", ""])
    for label, sc in [("85-100 exact (Conditions match)", 92),
                      ("60-84 probable (Symptoms/Title)", 72),
                      ("30-59 component-level only", 45),
                      ("1-29 weak lexical", 15)]:
        ws.append([label, ""])
        cell = ws.cell(row=ws.max_row, column=1)
        cell.fill = fill_for(sc)
        cell.font = Font(bold=True,
                         color="FFFFFF" if sc >= 60 else "202124")
    ws.column_dimensions["A"].width = 42; ws.column_dimensions["B"].width = 40
    for col, w in (("C", 34), ("D", 26), ("E", 22), ("F", 26), ("G", 22)):
        ws.column_dimensions[col].width = w

    def match_sheet(name: str, ms: List[Match]):
        ws = wb.create_sheet(name)
        sheet(ws, ["Score", "Band", "Bug ID", "Target Release", "Config Feature",
                   "Instances", "Item Title/Description", "Component",
                   "Severity/CVE", "Match Rationale", "Links", "Fixed In"],
              [7, 11, 12, 12, 30, 44, 60, 24, 18, 50, 34, 20])
        for m in ms:
            it = m.item
            ws.append([m.score, m.band, it.item_id, m.release, m.feature.label,
                       "; ".join(m.feature.instances),
                       it.title or it.description, it.component or it.category,
                       it.severity or " ".join(it.cves), m.reason,
                       ", ".join(f"{l}: {u}" for l, u in it.links),
                       ", ".join(it.fixed_versions)])
            c = ws.cell(row=ws.max_row, column=1)
            c.fill = fill_for(m.score)
            c.font = Font(bold=True,
                          color="FFFFFF" if m.score >= 60 else "202124")
            ws.cell(row=ws.max_row, column=7).alignment = Alignment(wrap_text=True)
        ws.auto_filter.ref = ws.dimensions

    # --- HeatMap sheet ---
    heat_rows, heat_max = compute_heatmap(matches, releases)
    ws = wb.create_sheet("HeatMap")
    rel_vs = [r.version for r in releases]
    sheet(ws, ["Config Feature (most -> least susceptible)", "Instances"]
          + [f"Susceptibility {v}" for v in rel_vs] + ["Total"],
          [42, 10] + [20] * len(rel_vs) + [12])
    for r in heat_rows:
        ws.append([r["label"], r["instances"]]
                  + [round(r["cells"][v]["index"], 2) for v in rel_vs]
                  + [round(r["total"], 2)])
        for i, v in enumerate(rel_vs):
            bg, fg = _heat_color(r["cells"][v]["index"], heat_max)
            cell = ws.cell(row=ws.max_row, column=3 + i)
            cell.fill = PatternFill("solid", fgColor=bg.lstrip("#").upper())
            cell.font = Font(bold=True, color=fg.lstrip("#").upper())
            cell.alignment = Alignment(horizontal="center")
    ws.append([])
    ws.append(["Index = weighted matched RN items remaining PRESENT after "
               "upgrading to that release (known issues on target branch + "
               "config-matched bugs not fixed by target). "
               "Weights: exact x3, probable x2, component x1, weak x0.25. "
               "Color normalized across matrix: green = least susceptible, "
               "red = most."])
    ws.cell(row=ws.max_row, column=1).font = Font(italic=True, size=9)
    ws.freeze_panes = "C2"

    fix_ms = [m for m in matches if not m.item.known_issue]
    risk_ms = [m for m in matches if m.item.known_issue]
    match_sheet("Associations", fix_ms)          # consolidated, filterable
    match_sheet("UpgradeRisks", risk_ms)         # consolidated, filterable
    for rel in releases:                         # per-release breakouts
        rv = rel.version
        match_sheet(f"Assoc {rv}"[:31],
                    [m for m in fix_ms if rv in m.release.split(", ")])
        match_sheet(f"Risks {rv}"[:31],
                    [m for m in risk_ms if rv in m.release.split(", ")])

    ws = wb.create_sheet("PlatformCVEs")
    sheet(ws, ["Bug ID", "Target Release", "CVE", "Item", "Links", "Fixed In"],
          [12, 14, 34, 70, 40, 22])
    for rv, it in sorted(platform, key=lambda x: x[1].item_id):
        ws.append([it.item_id, rv, " ".join(it.cves), it.title or it.description,
                   ", ".join(f"{l}: {u}" for l, u in it.links),
                   ", ".join(it.fixed_versions)])
    ws.auto_filter.ref = ws.dimensions

    ws = wb.create_sheet("AlreadyRemediated")
    sheet(ws, ["Bug ID", "Item", "Fixed In"], [12, 90, 26])
    for rv, it in sorted(remediated, key=lambda x: x[1].item_id):
        ws.append([it.item_id, it.title or it.description,
                   ", ".join(it.fixed_versions)])
    ws.auto_filter.ref = ws.dimensions

    ws = wb.create_sheet("Unmatched")
    sheet(ws, ["Target Release", "Bug ID", "Known Issue", "Component",
               "Severity/CVE", "Item"], [14, 12, 12, 26, 18, 90])
    for rv, items in unmatched.items():
        for it in items:
            ws.append([rv, it.item_id, "yes" if it.known_issue else "",
                       it.component or it.category,
                       it.severity or " ".join(it.cves),
                       it.title or it.description])
    ws.auto_filter.ref = ws.dimensions

    wb.save(path)


# =============================================================================
# CLI
# =============================================================================

def main(argv: Optional[List[str]] = None) -> int:
    ap = argparse.ArgumentParser(
        description="Associate BIG-IP QKView configuration with TMOS release-note "
                    "items and produce a scored HTML + Excel advisory report.")
    src = ap.add_mutually_exclusive_group(required=True)
    src.add_argument("--qkview-id", help="iHealth QKView ID (live pull via Files API)")
    src.add_argument("--offline", action="store_true",
                     help="read local config files instead of iHealth")
    ap.add_argument("--bigip-conf", help="path to bigip.conf (offline mode)")
    ap.add_argument("--gtm-conf", help="path to bigip_gtm.conf (offline mode)")
    ap.add_argument("--base-conf", help="path to bigip_base.conf (optional)")
    ap.add_argument("--rn", action="append", required=True,
                    help="TMOS release-note HTML file (repeatable)")
    ap.add_argument("--current-version", default="17.1.3",
                    help="running TMOS version on the device (default 17.1.3)")
    ap.add_argument("--hostname", default="", help="device label for the report")
    ap.add_argument("--out-prefix", default="tmos_rn_report",
                    help="output file prefix (default tmos_rn_report)")
    args = ap.parse_args(argv)

    # ---- acquire configuration
    conf_texts: Dict[str, str] = {}
    hostname = args.hostname
    if args.offline:
        if not args.bigip_conf:
            ap.error("--offline requires --bigip-conf")
        for label, p in (("bigip.conf", args.bigip_conf),
                         ("gtm.conf", args.gtm_conf),
                         ("bigip_base.conf", args.base_conf)):
            if p:
                if not os.path.isfile(p):
                    print(f"error: {p} not found", file=sys.stderr); return 2
                conf_texts[label] = open(p, encoding="utf-8",
                                         errors="replace").read()
        hostname = hostname or "offline-config"
    else:
        print(f"Pulling config for QKView {args.qkview_id} from iHealth ...",
              file=sys.stderr)
        conf_texts, ident = fetch_config_from_ihealth(args.qkview_id)
        hostname = hostname or ident.get("hostname", f"qkview-{args.qkview_id}")
        if ident.get("version") and args.current_version == "17.1.3":
            args.current_version = ident["version"]
        if not conf_texts:
            print("error: no config files retrieved from iHealth", file=sys.stderr)
            return 2

    # ---- parse config, detect features
    objs: List[ConfigObject] = []
    for label, text in conf_texts.items():
        parsed = parse_tmsh(text)
        print(f"  parsed {label}: {len(parsed)} stanzas", file=sys.stderr)
        objs.extend(parsed)
    features, comps, prov_info = detect_features(objs)
    if prov_info["provision"]:
        active = {m: l for m, l in prov_info["provision"].items()
                  if l.lower() != "none"}
        print(f"  provisioned modules: "
              f"{', '.join(f'{m}={l}' for m, l in sorted(active.items()))}",
              file=sys.stderr)
    print(f"  detected {len(features)} active features, "
          f"{sum(len(f.instances) for f in features)} instances", file=sys.stderr)

    # ---- parse release notes
    releases: List[ReleaseNotes] = []
    for p in args.rn:
        if not os.path.isfile(p):
            print(f"error: release-note file {p} not found", file=sys.stderr)
            return 2
        rn = parse_release_notes(p)
        print(f"  parsed {rn.filename}: v{rn.version} build {rn.build}, "
              f"{len(rn.items)} items "
              f"({sum(1 for i in rn.items.values() if i.known_issue)} known issues)",
              file=sys.stderr)
        releases.append(rn)
    releases.sort(key=lambda r: _vtuple(r.version))

    # ---- match + score
    matches, platform, remediated, unmatched = run_matching(
        features, releases, comps, args.current_version)
    print(f"  matches: {len(matches)} "
          f"(exact {sum(1 for m in matches if m.band=='exact')}, "
          f"probable {sum(1 for m in matches if m.band=='probable')}, "
          f"component {sum(1 for m in matches if m.band=='component')}, "
          f"weak {sum(1 for m in matches if m.band=='weak')}); "
          f"platform CVEs {len(platform)}; remediated {len(remediated)}",
          file=sys.stderr)

    # ---- release comparison / recommendation
    comparison = compare_releases(matches, platform, releases)
    for r in comparison:
        print(f"  {r['version']}: fixed={r['fixed']} residual={r['residual']} "
              f"risks={r['risks']} platform_fixed={r['platform_fixed']} "
              f"risk_score={r['risk_score']}"
              f"{'  <== RECOMMENDED' if r['recommended'] else ''}",
              file=sys.stderr)

    # ---- reports
    html_path = f"{args.out_prefix}.html"
    xlsx_path = f"{args.out_prefix}.xlsx"
    with open(html_path, "w", encoding="utf-8") as fh:
        fh.write(build_html(features, releases, matches, platform, remediated,
                            unmatched, comps, args.current_version, hostname,
                            comparison))
    build_xlsx(xlsx_path, features, releases, matches, platform, remediated,
               unmatched, args.current_version, hostname, comparison)
    print(f"wrote {html_path}\nwrote {xlsx_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
